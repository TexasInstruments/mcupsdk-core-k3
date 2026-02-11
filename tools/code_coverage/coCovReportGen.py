import csv, os, shutil
import openpyxl
import math
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
import argparse



parser=argparse.ArgumentParser(
    description=''' ## SDL Code Coverage Report Generation ##''',
    epilog="""Use this script to generate code coverage summary and gap report for SDL""")

parser.add_argument('-mod', '--moduleName',type=str,help='Module Name')
parser.add_argument('-csv', '--csv',type=str,help='csv file name')
#parser.add_argument('-sdl', '--sdlProjectPath',type=str,help='sdl project path')
parser.add_argument('-dev', '--deviceName', type=str, help='Device Name')
parser.add_argument('-rep', '--consolidatedReportPath', type=str, help='Consolidated Report Path')
args=parser.parse_args()


MODULE_NAME = args.moduleName


MODULE_CSV_FILE_NAME = args.csv


#PROJECT_PATH = args.sdlProjectPath

TARGET_DEVICE = args.deviceName

CONSOLIDATED_REPORT_PATH = args.consolidatedReportPath
# CONSOLIDATED_REPORT_PATH=os.getcwd()
COVERAGE_REPORT_FILE = CONSOLIDATED_REPORT_PATH + "/" + args.deviceName + "/" + MODULE_CSV_FILE_NAME
CONSOLIDATED_REPORT_FILE=CONSOLIDATED_REPORT_PATH + "/" + args.deviceName + "/" + MODULE_CSV_FILE_NAME + ".xlsx"

TEMPLATE_FILE = CONSOLIDATED_REPORT_PATH+"/ReportTemplate/DynamicAnalysisReportTemplate.xlsx"
CONSOLIDATED_REPORT_TEMPLATE = CONSOLIDATED_REPORT_PATH+ "/" + args.deviceName + "/DynamicAnalysisReportTemplate.xlsx"

COVERAGE_HTML_PATH=CONSOLIDATED_REPORT_PATH + "/" + args.deviceName + "/html"

if os.path.isfile(COVERAGE_REPORT_FILE) == False:
    print("[ ERROR ] "+COVERAGE_REPORT_FILE+" not found, Exiting")
    exit()

if os.path.isfile(TEMPLATE_FILE) == False:
    print("[ ERROR ] "+TEMPLATE_FILE+" not found, Exiting")
    exit()

if os.path.isfile(CONSOLIDATED_REPORT_TEMPLATE):
    os.remove(CONSOLIDATED_REPORT_TEMPLATE)
shutil.copy(TEMPLATE_FILE, CONSOLIDATED_REPORT_PATH + "/" + args.deviceName)


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

def formatCells(cellType):
    cellType.border = thin_border
    cellType.font = Font(name='Calibri', size=11)

wbFile = openpyxl.load_workbook(CONSOLIDATED_REPORT_TEMPLATE)

def getModuleName():
    if MODULE_NAME == "tog":
        moduleName = "stog"
    else :
        moduleName = MODULE_NAME
    return moduleName



#
# Populate Covered File List Sheet
#
csvFile = open(COVERAGE_REPORT_FILE, 'rt')
csvRawData = csv.reader(csvFile)
coverageFileListSheet = wbFile['Covered File List']
startListing = 0
index = 5
prevMod='None'
statementCovDict=dict()
branchCovDict=dict()
functionCovDict=dict()
mcdcCovDict=dict()
moduleList=set()
sModIndex = 1
bModIndex = 1
mModIndex = 1
fModIndex = 1
mcdcNAtag=0
filenameLocDict=dict()
for row in csvRawData:
    if row and (startListing == 0):
        if row[0].find("Coverage File List") != -1:
            startListing = 1
            continue
    if startListing == 1:
        if row and any(cell.strip() for cell in row):
            if row[0].find("S No.") != -1:
                continue
            cellslno = coverageFileListSheet.cell(row=index, column=2)
            cellslno.value = index-4
            formatCells(cellslno)
            cellFileName = coverageFileListSheet.cell(row=index, column=3)
            cellFileName.value = row[1]
            fileName = row[1]
            while fileName.find("/") != -1:
                fileName = fileName[fileName.find("/")+1:]
            filenameLocDict.setdefault(fileName,{})["row"] = index
            formatCells(cellFileName)
            print("Current filename: ",fileName)

            totalStatements = coverageFileListSheet.cell(row=index, column=4)
            if int(row[2]) != 0:
                totalStatements.value = int(row[2])
            else:
                totalStatements.value = 0
            formatCells(totalStatements)

            executedStatements = coverageFileListSheet.cell(row=index, column=5)
            if int(row[3]) != 0:
                executedStatements.value = int(row[3])
            else:
                executedStatements.value = 0
            formatCells(executedStatements)

            totalBranches = coverageFileListSheet.cell(row=index, column=6)
            if int(row[4]) != 0:
                totalBranches.value = int(row[4])
            else:
                totalBranches.value = 0
            formatCells(totalBranches)

            executedBranches = coverageFileListSheet.cell(row=index, column=7)
            if int(row[5]) != 0:
                executedBranches.value = int(row[5])
            else:
                executedBranches.value = 0
            formatCells(executedBranches)

            totalMCDC = coverageFileListSheet.cell(row=index, column=8)
            if int(row[6]) != 0:
                totalMCDC.value = int(row[6])
            else:
                totalMCDC.value = "NA"
                totalMCDC.alignment = Alignment(horizontal='right')
            formatCells(totalMCDC)

            executedMCDC = coverageFileListSheet.cell(row=index, column=9)
            if int(row[7]) != 0:
                executedMCDC.value = int(row[7])
            else:
                executedMCDC.value = "NA"
                executedMCDC.alignment = Alignment(horizontal='right')
            formatCells(executedMCDC)

            formatCells(coverageFileListSheet.cell(row=index, column=10))

            modName = getModuleName()
            print("Current Modname:", modName)
            totalStatements = int(row[2])
            totalStatementsExecuted = int(row[3])
            totalBranch = int(row[4])
            totalBranchExecuted = int(row[5])
            totalmcdc = int(row[6])
            Executedmcdc = int(row[7])
            totalFunctions = int(row[8])
            ExecutedFunctions = int(row[9])

            #Handling NA MCDC Coverage
            if totalmcdc == 0:
                mcdcNAtag = 1

            if prevMod != modName:
                if sModIndex != 1:
                    statementCovDict[prevMod] = round(statementCovDict[prevMod]/sModIndex,2)
                if bModIndex != 1:
                    branchCovDict[prevMod] = round(branchCovDict[prevMod]/bModIndex,2)
                if mModIndex != 1:
                    mcdcCovDict[prevMod] = round(mcdcCovDict[prevMod]/mModIndex,2)
                if fModIndex != 1:
                    functionCovDict[prevMod] = round(functionCovDict[prevMod]/fModIndex,2)
                sModIndex = 1
                bModIndex = 1
                mModIndex = 1
                fModIndex = 1
                if totalStatements != 0:
                    statementCovDict[modName] = round((totalStatementsExecuted/totalStatements)*100,2)
                if totalBranch != 0:
                    branchCovDict[modName] = round((totalBranchExecuted/totalBranch)*100,2)
                if totalmcdc != 0:
                    mcdcCovDict[modName] = round((Executedmcdc/totalmcdc)*100,2)
                if totalFunctions != 0:
                    functionCovDict[modName] = round((ExecutedFunctions/totalFunctions)*100,2)
                prevMod = modName
                moduleList.add(modName)
            else:
                if totalStatements != 0:
                    if modName not in statementCovDict:
                        statementCovDict[modName] = 0
                        sModIndex-=1
                    statementCovDict[modName] += round((totalStatementsExecuted/totalStatements)*100,2)
                    sModIndex+=1

                if totalBranch != 0:
                    if modName not in branchCovDict:
                        branchCovDict[modName] = 0
                        bModIndex-=1
                    branchCovDict[modName] += round((totalBranchExecuted/totalBranch)*100,2)
                    bModIndex+=1

                if totalmcdc != 0:
                    if modName not in mcdcCovDict or mcdcNAtag == 1:
                        mcdcCovDict[modName] = round((Executedmcdc/totalmcdc)*100,2)
                        mcdcNAtag = 0
                        mModIndex-=1
                    else:
                        mcdcCovDict[modName] += round((Executedmcdc/totalmcdc)*100,2)
                    mModIndex+=1

                if totalFunctions != 0:
                    if modName not in functionCovDict:
                        functionCovDict[modName] = 0
                        fModIndex-=1
                    functionCovDict[modName] += round((ExecutedFunctions/totalFunctions)*100,2)
                    fModIndex+=1

            index+=1
        else:
            break
if branchCovDict:
    branchCovDict[prevMod] = round(branchCovDict[prevMod]/bModIndex,2)
if statementCovDict:
    statementCovDict[prevMod] = round(statementCovDict[prevMod]/sModIndex,2)
if mcdcCovDict:
    mcdcCovDict[prevMod] = round(mcdcCovDict[prevMod]/mModIndex,2)
if functionCovDict:
    functionCovDict[prevMod] = round(functionCovDict[prevMod]/fModIndex,2)
csvFile.close()

#
# Populate Summary Sheet
#
coverageSummarySheet = wbFile['Summary']
csvFile = open(COVERAGE_REPORT_FILE, 'rt')
csvRawData = csv.reader(csvFile)
for row in csvRawData:
    if row:
        if row[0].find("Total number of files") != -1:
            coverageSummarySheet['D13'] = int(row[1])
        if row[0].find("Total number of functions") != -1:
            coverageSummarySheet['D14'] = int(row[1])
        if row[0].find("Percentage of functions covered") != -1:
            coverageSummarySheet['D16'] = float(row[1][:-1])/100
            coverageSummarySheet['D16'].number_format = '0.00%'
        if row[0].find("Average Coverage Achieved") != -1:
            coverageSummarySheet['D20'] = float(row[1][:-1])/100
            coverageSummarySheet['D20'].number_format = '0.00%'
            coverageSummarySheet['E20'] = float(row[2][:-1])/100
            coverageSummarySheet['E20'].number_format = '0.00%'
            coverageSummarySheet['G20'] = float(row[4][:-1])/100
            coverageSummarySheet['G20'].number_format = '0.00%'
            coverageSummarySheet['F20'] = str(row[3]).upper()
        if row[0].find("Number of functions NOT covered") != -1:
            coverageSummarySheet['D15'] = int(row[1])

index = 0
for module in moduleList:
    moduleCellInSummary = coverageSummarySheet.cell(row=24+index, column=3)
    moduleCellInSummary.value = module
    formatCells(moduleCellInSummary)
    moduleCellInSummary.alignment = Alignment(horizontal='center')
    moduleCellInSummary.font = Font(bold=True)
    moduleCellInSummary.fill = PatternFill(start_color='2BF6F6',
                   end_color='2BF6F6',
                   fill_type='solid')

    if statementCovDict:
        statementCovCellInSummary = coverageSummarySheet.cell(row=24+index, column=4)
        statementCovCellInSummary.value = float(statementCovDict[module])/100
        formatCells(statementCovCellInSummary)
        statementCovCellInSummary.number_format = '0.00%'
        statementCovCellInSummary.alignment = Alignment(horizontal='center')

        branchCovCellInSummary = coverageSummarySheet.cell(row=24+index, column=5)
        if branchCovDict:
            branchCovCellInSummary.value = float(branchCovDict[module])/100
        else:
            branchCovCellInSummary.value = 'NA'
        branchCovCellInSummary.number_format = '0.00%'
        formatCells(branchCovCellInSummary)
        branchCovCellInSummary.alignment = Alignment(horizontal='center')

        mcdcCovCellInSummary = coverageSummarySheet.cell(row=24+index, column=6)
        if mcdcCovDict:
            mcdcCovCellInSummary.value = float(mcdcCovDict[module])/100
        else:
            mcdcCovCellInSummary.value = 'NA'
        formatCells(mcdcCovCellInSummary)
        mcdcCovCellInSummary.alignment = Alignment(horizontal='center')

    if functionCovDict:
        fnCallCovCellInSummary = coverageSummarySheet.cell(row=24+index, column=7)
        fnCallCovCellInSummary.value = float(functionCovDict[module])/100
        fnCallCovCellInSummary.number_format = '0.00%'
        formatCells(fnCallCovCellInSummary)
        fnCallCovCellInSummary.alignment = Alignment(horizontal='center')
    index += 1
csvFile.close()

#
# GAP Reason html Parsing
# ***************************************************************************************************************************
#
COVERAGE_COVEREDFILE_COMMENT_TAG ="TI_COVERAGE_FILE_COMMENT"
COVERAGE_GAP_START_TAG ="TI_COVERAGE_GAP_START"
COVERAGE_GAP_STOP_TAG ="TI_COVERAGE_GAP_STOP"
HTML_LINE_TAG="href='#L"
gapDict={}
CoveredFileCommentsDict=dict()
gapIndex = 0



for module in moduleList:
    moduleHtmlSrcDirectory=COVERAGE_HTML_PATH+"/"+module
    for root, dirs, files in os.walk(moduleHtmlSrcDirectory):
        for filename in files:
            if filename.endswith('.html'):
                with open(os.path.join(root, filename)) as htmlFile:
                    htmlFile_data = htmlFile.readlines()
                    coverageEndTagSearch = False
                    for rawLine in htmlFile_data:
                        line = rawLine
                        if (line.find(COVERAGE_COVEREDFILE_COMMENT_TAG) != -1):
                            line = line[line.find(COVERAGE_COVEREDFILE_COMMENT_TAG)+len(COVERAGE_COVEREDFILE_COMMENT_TAG):]
                            CoveredFileCommentsDict[filename[:-5]] = line[:line.find("*/")]

                    for rawLine in htmlFile_data:
                        line = rawLine
                        while 1:
                            if (line.find(COVERAGE_GAP_START_TAG) != -1) and (coverageEndTagSearch == False):
                                tempLine = line
                                while len(tempLine) > 8:
                                    StartTagLoc = tempLine.find(COVERAGE_GAP_START_TAG)
                                    htmlTagLoc = tempLine.find(HTML_LINE_TAG)
                                    if htmlTagLoc < StartTagLoc:
                                        tempLine = tempLine[len(HTML_LINE_TAG)+htmlTagLoc:]
                                    if htmlTagLoc > StartTagLoc:
                                        gapDict.setdefault(gapIndex,{})["FILE_NAME"] = filename[:-5]
                                        gapDict.setdefault(gapIndex,{})["L_START"] = tempLine[tempLine.find("<pre>")+5:tempLine.find("</pre>")]
                                        coverageEndTagSearch = True
                                        break
                                gapDict.setdefault(gapIndex,{})["RATIONALE"]  = tempLine[tempLine.find(COVERAGE_GAP_START_TAG)+len(COVERAGE_GAP_START_TAG):tempLine.find("*/")]
                            if (line.find(COVERAGE_GAP_STOP_TAG) != -1) and (coverageEndTagSearch == True):
                                tempLine = line
                                while len(tempLine) > 8:
                                    stopTagLoc = tempLine.find(COVERAGE_GAP_STOP_TAG)
                                    htmlTagLoc = tempLine.find(HTML_LINE_TAG)
                                    if htmlTagLoc < stopTagLoc:
                                        tempLine = tempLine[len(HTML_LINE_TAG)+htmlTagLoc:]
                                    if htmlTagLoc > stopTagLoc:
                                        gapDict.setdefault(gapIndex,{})["L_STOP"] = tempLine[tempLine.find("<pre>")+5:tempLine.find("</pre>")]
                                        gapIndex+=1
                                        coverageEndTagSearch = False
                                        break
                            line = line[line.find(COVERAGE_GAP_STOP_TAG)+len(COVERAGE_GAP_STOP_TAG):]
                            if (line.find(COVERAGE_GAP_STOP_TAG) == -1) and (line.find(COVERAGE_GAP_START_TAG) == -1):
                                break


#
# Check for duplicates in dictionary
#
removeList = set()
for SearchMod in gapDict:
    for recurMod in list(gapDict)[SearchMod+1:len(list(gapDict))]:
        if (gapDict[recurMod]["FILE_NAME"] == gapDict[SearchMod]["FILE_NAME"]) and (gapDict[recurMod]["L_START"] == gapDict[SearchMod]["L_START"]) and (gapDict[recurMod]["L_STOP"] == gapDict[SearchMod]["L_STOP"]):
            removeList.add(int(recurMod))
for i in removeList:
    gapDict.pop(i)

#
# GAP Reason Populate into Consolidated report
#
coverageSummarySheet = wbFile['Gap Report']
gapReportIndex = 0
for data in gapDict:
    gapCell = coverageSummarySheet.cell(row=8+gapReportIndex, column=3)
    gapCell.value = gapDict[data]["FILE_NAME"]
    gapCell.alignment = Alignment(vertical='top', horizontal="left")
    formatCells(gapCell)
    gapCell = coverageSummarySheet.cell(row=8+gapReportIndex, column=4)
    if (int(gapDict[data]["L_STOP"]) - int(gapDict[data]["L_START"])) == 2:
        gapCell.value = int(int(gapDict[data]["L_START"]) + 1)
    else:
        gapCell.value = str(int(gapDict[data]["L_START"])+1) +"-"+ str(int(gapDict[data]["L_STOP"])-1)
    gapCell.alignment = Alignment(vertical='top', horizontal="left")
    formatCells(gapCell)
    gapCell = coverageSummarySheet.cell(row=8+gapReportIndex, column=5)
    gapCell.value = gapDict[data]["RATIONALE"]
    formatCells(gapCell)
    gapCell.alignment = Alignment(wrap_text=True,vertical='top')
    gapCell = coverageSummarySheet.cell(row=8+gapReportIndex, column=6)
    if gapDict[data]["RATIONALE"].find("Branch Gap") != -1:
        gapCell.value = "100% branch coverage cannot be achieved"
    else:
        gapCell.value = "100% statement coverage cannot be achieved"
    formatCells(gapCell)
    gapCell.alignment = Alignment(wrap_text=True,vertical='top')
    gapReportIndex+=1


#
# Populate Covered File List Sheet
#
# print(CoveredFileCommentsDict)
# print("")
# print(filenameLocDict)
# print("")
csvFile = open(COVERAGE_REPORT_FILE, 'rt')
csvRawData = csv.reader(csvFile)
coverageFileListSheet = wbFile['Covered File List']
for file in CoveredFileCommentsDict:
    cell = coverageFileListSheet.cell(row=filenameLocDict[file]['row'], column=10)
    cell.value = CoveredFileCommentsDict[file]
    cell.alignment = Alignment(wrap_text=True,horizontal='left')

wbFile.save(CONSOLIDATED_REPORT_TEMPLATE)

shutil.copy(CONSOLIDATED_REPORT_TEMPLATE, os.path.join(CONSOLIDATED_REPORT_PATH, args.deviceName, f"{MODULE_CSV_FILE_NAME}_CoverageReport.xlsx"))
os.remove(CONSOLIDATED_REPORT_TEMPLATE)
temp = os.getcwd()
os.chdir(os.path.join(CONSOLIDATED_REPORT_PATH,args.deviceName))
reportDir = os.getcwd()
os.chdir(temp)
print("#")
print("# Generated Consolidated Dynamic Analysis report at "+reportDir + "/" + MODULE_NAME  + "_CoverageReport.xlsx")
print("#")


